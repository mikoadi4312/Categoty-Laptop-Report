import argparse
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config.database import get_engine  # noqa: E402


REPORT_COLUMNS = [
    "BRAND",
    "DAY Qty",
    "DAY Rev (IDR)",
    "MTD Qty",
    "MTD Rev (IDR)",
    "SP Qty",
    "SP Rev",
    "New",
    "Demo",
    "Stock Day",
    "GROWTH RATE",
]

THEME_NAVY = "0B2F66"
THEME_BLUE = "2F6FED"
THEME_BLUE_DARK = "123F84"
THEME_SOFT = "F4F9FF"
THEME_LINE = "CFE0F5"
THEME_HEADER = "EFF4FF"
THEME_TOTAL = "DCEBFF"
THEME_GREEN = "0F8F5F"
THEME_INK = "000000"
THEME_MUTED = "475569"

THIN_SIDE = Side(style="thin", color=THEME_LINE)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MEDIUM_BOTTOM_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=Side(style="medium", color=THEME_NAVY),
)


def report_font(
    *,
    bold: bool = False,
    color: str = THEME_INK,
    size: int = 10,
) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def set_print_layout(ws, max_col: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{ws.max_row}"


def style_cell(
    cell,
    *,
    fill: str | None = None,
    font: Font | None = None,
    align: Alignment | None = None,
    border: Border = THIN_BORDER,
) -> None:
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    cell.border = border


def parse_date(value: str | None) -> date:
    if not value:
        return date.today()
    return datetime.strptime(value, "%Y-%m-%d").date()


def money(value) -> float:
    if pd.isna(value):
        return 0.0
    return float(value)


def rounded_integer(value):
    if value is None or pd.isna(value) or isinstance(value, bool):
        return value
    try:
        return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return value


def query_dataframe(
    engine,
    sql: str,
    target_date: date,
    sales_date: date | None = None,
    stock_date: date | None = None,
) -> pd.DataFrame:
    params = {
        "target_date": target_date,
        "sales_date": sales_date or target_date,
        "stock_date": stock_date or target_date,
    }
    with engine.connect() as conn:
        return pd.read_sql_query(text(sql), conn, params=params)


def build_laptop_report_df(engine, target_date: date, stock_date: date | None = None) -> pd.DataFrame:
    sales_date = target_date
    selected_stock_date = stock_date or target_date
    day_sql = """
        SELECT b.brand, SUM(s.day_qty) AS day_qty, SUM(s.day_revenue) AS day_rev
        FROM fact_sales s
        JOIN dim_brand b ON s.brand_id = b.id
        WHERE s.sale_date = :sales_date
        GROUP BY b.brand
    """
    mtd_sql = """
        SELECT b.brand, SUM(s.day_qty) AS mtd_qty, SUM(s.day_revenue) AS mtd_rev
        FROM fact_sales s
        JOIN dim_brand b ON s.brand_id = b.id
        WHERE DATE_TRUNC('month', s.sale_date) = DATE_TRUNC('month', CAST(:sales_date AS DATE))
          AND s.sale_date <= :sales_date
        GROUP BY b.brand
    """
    same_period_sql = """
        SELECT b.brand, SUM(s.day_qty) AS sp_qty, SUM(s.day_revenue) AS sp_rev
        FROM fact_sales s
        JOIN dim_brand b ON s.brand_id = b.id
        WHERE s.sale_date >= (DATE_TRUNC('month', CAST(:sales_date AS DATE)) - INTERVAL '1 month')::date
          AND s.sale_date <= LEAST(
              (DATE_TRUNC('month', CAST(:sales_date AS DATE)) - INTERVAL '1 day')::date,
              (
                  (DATE_TRUNC('month', CAST(:sales_date AS DATE)) - INTERVAL '1 month')
                  + ((EXTRACT(DAY FROM CAST(:sales_date AS DATE))::int - 1) * INTERVAL '1 day')
              )::date
          )
        GROUP BY b.brand
    """
    stock_sql = """
        SELECT b.brand,
               SUM(k.new_stock) AS new_stock,
               SUM(k.demo_units) AS demo_units,
               SUM(k.stock_volume) AS stock_volume
        FROM fact_stock k
        JOIN dim_brand b ON k.brand_id = b.id
        WHERE k.stock_date = :stock_date
        GROUP BY b.brand
    """

    day_df = query_dataframe(engine, day_sql, target_date, sales_date)
    mtd_df = query_dataframe(engine, mtd_sql, target_date, sales_date)
    same_period_df = query_dataframe(engine, same_period_sql, target_date, sales_date)
    stock_df = query_dataframe(engine, stock_sql, target_date, sales_date, selected_stock_date)

    current_laptop_brands = (
        set(day_df.get("brand", []))
        | set(mtd_df.get("brand", []))
        | set(stock_df.get("brand", []))
    )
    brands = pd.DataFrame({"brand": sorted(current_laptop_brands)})
    if brands.empty:
        columns = REPORT_COLUMNS
        return pd.DataFrame([build_total_row(pd.DataFrame(columns=columns), sales_date)], columns=columns)

    df = brands.merge(day_df, on="brand", how="left")
    df = df.merge(mtd_df, on="brand", how="left")
    df = df.merge(same_period_df, on="brand", how="left")
    df = df.merge(stock_df, on="brand", how="left")
    for col in [
        "day_qty",
        "day_rev",
        "mtd_qty",
        "mtd_rev",
        "sp_qty",
        "sp_rev",
        "new_stock",
        "demo_units",
        "stock_volume",
    ]:
        df[col] = df[col].fillna(0)

    days_elapsed = sales_date.day
    df["daily_rate"] = df["mtd_qty"] / days_elapsed
    df["stock_day"] = df.apply(
        lambda row: 0 if row["daily_rate"] == 0 else row["stock_volume"] / row["daily_rate"],
        axis=1,
    )
    df["growth_rate"] = df.apply(
        lambda row: 0 if row["sp_rev"] == 0 else (row["mtd_rev"] - row["sp_rev"]) / row["sp_rev"],
        axis=1,
    )

    df = df.sort_values(["mtd_rev", "brand"], ascending=[False, True])
    output = df.rename(
        columns={
            "brand": "BRAND",
            "day_qty": "DAY Qty",
            "day_rev": "DAY Rev (IDR)",
            "mtd_qty": "MTD Qty",
            "mtd_rev": "MTD Rev (IDR)",
            "sp_qty": "SP Qty",
            "sp_rev": "SP Rev",
            "new_stock": "New",
            "demo_units": "Demo",
            "stock_day": "Stock Day",
            "growth_rate": "GROWTH RATE",
        }
    )[REPORT_COLUMNS]

    total = build_total_row(output, sales_date)
    output = pd.concat([output, pd.DataFrame([total])], ignore_index=True)
    return output


def build_total_row(df: pd.DataFrame, target_date: date) -> dict:
    day_qty = int(df["DAY Qty"].sum()) if "DAY Qty" in df else 0
    day_rev = money(df["DAY Rev (IDR)"].sum()) if "DAY Rev (IDR)" in df else 0.0
    mtd_qty = int(df["MTD Qty"].sum()) if "MTD Qty" in df else 0
    mtd_rev = money(df["MTD Rev (IDR)"].sum()) if "MTD Rev (IDR)" in df else 0.0
    sp_qty = int(df["SP Qty"].sum()) if "SP Qty" in df else 0
    sp_rev = money(df["SP Rev"].sum()) if "SP Rev" in df else 0.0
    new_stock = int(df["New"].sum()) if "New" in df else 0
    demo_units = int(df["Demo"].sum()) if "Demo" in df else 0
    stock_volume = new_stock + demo_units
    daily_rate = mtd_qty / target_date.day if target_date.day else 0
    stock_day = 0 if daily_rate == 0 else stock_volume / daily_rate
    growth_rate = 0 if sp_rev == 0 else (mtd_rev - sp_rev) / sp_rev
    return {
        "BRAND": "GRAND TOTAL",
        "DAY Qty": day_qty,
        "DAY Rev (IDR)": day_rev,
        "MTD Qty": mtd_qty,
        "MTD Rev (IDR)": mtd_rev,
        "SP Qty": sp_qty,
        "SP Rev": sp_rev,
        "New": new_stock,
        "Demo": demo_units,
        "Stock Day": stock_day,
        "GROWTH RATE": growth_rate,
    }


def build_shop_level_df(engine, target_date: date) -> tuple[pd.DataFrame, list[str]]:
    sales_date = target_date
    day_sql = """
        SELECT st.id_store, st.store_name, b.brand,
               SUM(s.day_qty) AS day_qty, SUM(s.day_revenue) AS day_revenue
        FROM fact_sales s
        JOIN dim_store st ON s.store_id = st.id
        JOIN dim_brand b ON s.brand_id = b.id
        WHERE s.sale_date = :sales_date
        GROUP BY st.id_store, st.store_name, b.brand
    """
    mtd_sql = """
        SELECT st.id_store, st.store_name, b.brand,
               SUM(s.day_qty) AS mtd_qty, SUM(s.day_revenue) AS mtd_revenue
        FROM fact_sales s
        JOIN dim_store st ON s.store_id = st.id
        JOIN dim_brand b ON s.brand_id = b.id
        WHERE DATE_TRUNC('month', s.sale_date) = DATE_TRUNC('month', CAST(:sales_date AS DATE))
          AND s.sale_date <= :sales_date
        GROUP BY st.id_store, st.store_name, b.brand
    """
    day_df = query_dataframe(engine, day_sql, target_date, sales_date)
    mtd_df = query_dataframe(engine, mtd_sql, target_date, sales_date)

    brands = sorted(set(day_df.get("brand", [])) | set(mtd_df.get("brand", [])))
    if not brands:
        return pd.DataFrame(columns=["STORE NAME", "TOTAL DAY Rev", "TOTAL MTD Rev"]), []

    store_keys = pd.concat(
        [
            day_df[["id_store", "store_name"]] if not day_df.empty else pd.DataFrame(columns=["id_store", "store_name"]),
            mtd_df[["id_store", "store_name"]] if not mtd_df.empty else pd.DataFrame(columns=["id_store", "store_name"]),
        ],
        ignore_index=True,
    ).drop_duplicates()

    brand_order_df = (
        mtd_df.groupby("brand", as_index=False)["mtd_revenue"].sum().sort_values(["mtd_revenue", "brand"], ascending=[False, True])
        if not mtd_df.empty
        else pd.DataFrame({"brand": brands, "mtd_revenue": 0})
    )
    brands = brand_order_df["brand"].tolist()

    rows = []
    for _, store in store_keys.iterrows():
        row = {"STORE NAME": store["store_name"]}
        total_day = 0.0
        total_mtd = 0.0
        for brand in brands:
            day_value = day_df[
                (day_df["id_store"] == store["id_store"]) & (day_df["brand"] == brand)
            ]["day_revenue"].sum()
            mtd_value = mtd_df[
                (mtd_df["id_store"] == store["id_store"]) & (mtd_df["brand"] == brand)
            ]["mtd_revenue"].sum()
            row[f"{brand} DAY Rev"] = money(day_value)
            row[f"{brand} MTD Rev"] = money(mtd_value)
            total_day += money(day_value)
            total_mtd += money(mtd_value)

        row["TOTAL DAY Rev"] = total_day
        row["TOTAL MTD Rev"] = total_mtd
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.sort_values(["TOTAL MTD Rev", "STORE NAME"], ascending=[False, True])

    total_row = {"STORE NAME": "GRAND TOTAL"}
    for brand in brands:
        total_row[f"{brand} DAY Rev"] = money(df[f"{brand} DAY Rev"].sum())
        total_row[f"{brand} MTD Rev"] = money(df[f"{brand} MTD Rev"].sum())
    total_row["TOTAL DAY Rev"] = money(df["TOTAL DAY Rev"].sum())
    total_row["TOTAL MTD Rev"] = money(df["TOTAL MTD Rev"].sum())
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df, brands


def write_laptop_report_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    target_date: date,
    stock_date: date,
) -> None:
    ws = wb.active
    ws.title = "LAPTOP REPORT"

    max_col = len(REPORT_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(
        1,
        1,
        f"LAPTOP CATEGORY DAILY REPORT | SALES {target_date:%Y-%m-%d} | STOCK {stock_date:%Y-%m-%d}",
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 28

    ws.cell(2, 1, "")
    ws.merge_cells("B2:C2")
    ws.cell(2, 2, "DAY")
    ws.merge_cells("D2:E2")
    ws.cell(2, 4, "MTD")
    ws.merge_cells("F2:G2")
    ws.cell(2, 6, "SAME PERIOD")
    ws.merge_cells("H2:K2")
    ws.cell(2, 8, "STOCK")

    for col_idx, column in enumerate(REPORT_COLUMNS, start=1):
        ws.cell(3, col_idx, column)

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)

    for col_idx in range(1, max_col + 1):
        style_cell(
            ws.cell(1, col_idx),
            fill=THEME_NAVY,
            font=report_font(bold=True, color="FFFFFF", size=15),
            align=Alignment(horizontal="center", vertical="center"),
            border=MEDIUM_BOTTOM_BORDER,
        )

    group_colors = {
        1: THEME_NAVY,
        2: THEME_BLUE,
        3: THEME_BLUE,
        4: THEME_BLUE_DARK,
        5: THEME_BLUE_DARK,
        6: THEME_MUTED,
        7: THEME_MUTED,
        8: THEME_GREEN,
        9: THEME_GREEN,
        10: THEME_GREEN,
        11: THEME_GREEN,
    }
    for row_idx in [2, 3]:
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if row_idx == 2:
                style_cell(
                    cell,
                    fill=group_colors[col_idx],
                    font=report_font(bold=True, color="FFFFFF", size=10),
                    align=Alignment(horizontal="center", vertical="center"),
                    border=MEDIUM_BOTTOM_BORDER,
                )
            else:
                style_cell(
                    cell,
                    fill=THEME_HEADER,
                    font=report_font(bold=True, color=THEME_NAVY, size=10),
                    align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                    border=MEDIUM_BOTTOM_BORDER,
                )

    for row_idx in range(4, ws.max_row + 1):
        is_total_row = ws.cell(row_idx, 1).value == "GRAND TOTAL"
        fill = THEME_TOTAL if is_total_row else (THEME_SOFT if row_idx % 2 == 0 else "FFFFFF")
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            style_cell(
                cell,
                fill=fill,
                font=report_font(bold=is_total_row, color=THEME_INK, size=10),
                align=Alignment(
                    horizontal="left" if col_idx == 1 else "right",
                    vertical="center",
                    wrap_text=(col_idx == 1),
                ),
                border=MEDIUM_BOTTOM_BORDER if is_total_row else THIN_BORDER,
            )
        ws.row_dimensions[row_idx].height = 22

    for col_idx in [2, 4, 6, 8, 9, 10]:
        for row_idx in range(4, ws.max_row + 1):
            ws.cell(row_idx, col_idx).number_format = "#,##0"
    for col_idx in [3, 5, 7]:
        for row_idx in range(4, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = rounded_integer(cell.value)
            cell.number_format = "#,##0"
    for row_idx in range(4, ws.max_row + 1):
        ws.cell(row_idx, 11).number_format = "0.00%"

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 17
    ws.freeze_panes = "A4"
    set_print_layout(ws, max_col)


def write_shop_level_sheet(wb: Workbook, df: pd.DataFrame, brands: list[str], target_date: date) -> None:
    ws = wb.create_sheet("SHOP LEVEL")
    max_col = 1 + (len(brands) * 2) + 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1, f"LAPTOP SHOP LEVEL REPORT - {target_date:%Y-%m-%d}")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 28

    ws.cell(2, 1, "SHOP")
    start_col = 2
    for idx, brand in enumerate(brands):
        end_col = start_col + 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(2, start_col, brand)
        start_col += 2
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 1)
    ws.cell(2, start_col, "TOTAL")

    ws.cell(3, 1, "STORE NAME")
    col_idx = 2
    ordered_columns = ["STORE NAME"]
    for brand in brands:
        ws.cell(3, col_idx, "DAY Rev")
        ws.cell(3, col_idx + 1, "MTD Rev")
        ordered_columns.extend([f"{brand} DAY Rev", f"{brand} MTD Rev"])
        col_idx += 2
    ws.cell(3, col_idx, "DAY Rev")
    ws.cell(3, col_idx + 1, "MTD Rev")
    ordered_columns.extend(["TOTAL DAY Rev", "TOTAL MTD Rev"])

    for row_idx, row in enumerate(df[ordered_columns].itertuples(index=False), start=4):
        for cell_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, cell_idx, value)

    for col_idx in range(1, max_col + 1):
        style_cell(
            ws.cell(1, col_idx),
            fill=THEME_NAVY,
            font=report_font(bold=True, color="FFFFFF", size=15),
            align=Alignment(horizontal="center", vertical="center"),
            border=MEDIUM_BOTTOM_BORDER,
        )

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(2, col_idx)
        if col_idx == 1:
            group_fill = THEME_NAVY
        elif col_idx >= max_col - 1:
            group_fill = THEME_GREEN
        else:
            brand_group = (col_idx - 2) // 2
            group_fill = THEME_BLUE if brand_group % 2 == 0 else THEME_BLUE_DARK
        style_cell(
            cell,
            fill=group_fill,
            font=report_font(bold=True, color="FFFFFF", size=10),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=MEDIUM_BOTTOM_BORDER,
        )

        sub_cell = ws.cell(3, col_idx)
        style_cell(
            sub_cell,
            fill=THEME_HEADER,
            font=report_font(bold=True, color=THEME_NAVY, size=10),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=MEDIUM_BOTTOM_BORDER,
        )

    total_start_col = max_col - 1
    for row_idx in range(4, ws.max_row + 1):
        is_total_row = ws.cell(row_idx, 1).value == "GRAND TOTAL"
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            is_total_col = col_idx >= total_start_col
            if is_total_row:
                fill = THEME_TOTAL
                font = report_font(bold=True)
                border = MEDIUM_BOTTOM_BORDER
            elif is_total_col:
                fill = "EAF7F1"
                font = report_font(bold=True)
                border = THIN_BORDER
            else:
                fill = THEME_SOFT if row_idx % 2 == 0 else "FFFFFF"
                font = report_font()
                border = THIN_BORDER
            style_cell(
                cell,
                fill=fill,
                font=font,
                align=Alignment(
                    horizontal="left" if col_idx == 1 else "right",
                    vertical="center",
                    wrap_text=(col_idx == 1),
                ),
                border=border,
            )
        ws.row_dimensions[row_idx].height = 22

    for row_idx in range(4, ws.max_row + 1):
        for col_idx in range(2, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = rounded_integer(cell.value)
            cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 42
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.freeze_panes = "B4"
    set_print_layout(ws, max_col)


def generate_report(target_date: date, output_path: Path, stock_date: date | None = None) -> None:
    selected_stock_date = stock_date or target_date
    engine = get_engine()
    laptop_df = build_laptop_report_df(engine, target_date, selected_stock_date)
    shop_df, brands = build_shop_level_df(engine, target_date)

    wb = Workbook()
    write_laptop_report_sheet(wb, laptop_df, target_date, selected_stock_date)
    write_shop_level_sheet(wb, shop_df, brands, target_date)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate LAPTOP daily Excel report.")
    parser.add_argument("--date", help="Target date, YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--stock-date", help="Stock snapshot date, YYYY-MM-DD. Defaults to target date.")
    parser.add_argument("--output", help="Output filename. Defaults to LAPTOP_REPORT_YYYYMMDD.xlsx.")
    args = parser.parse_args()

    target_date = parse_date(args.date)
    stock_date = parse_date(args.stock_date) if args.stock_date else target_date
    output = args.output or f"LAPTOP_REPORT_{target_date:%Y%m%d}.xlsx"
    output_path = Path(output).expanduser().resolve()

    generate_report(target_date, output_path, stock_date)
    print(f"Report generated: {output_path}")


if __name__ == "__main__":
    main()
