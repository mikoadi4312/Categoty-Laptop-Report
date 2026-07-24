import sys
from datetime import date, datetime
from numbers import Number
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config.database import ensure_schema, get_connection  # noqa: E402
from etl.upload_sales import upload_sales  # noqa: E402
from etl.upload_stock import stock_date_from_filename, upload_stock  # noqa: E402
from reports.generate_laptop_report import generate_report  # noqa: E402


APP_DATA_DIR = PROJECT_ROOT / "app_data"
UPLOAD_DIR = APP_DATA_DIR / "uploads"
REPORT_DIR = APP_DATA_DIR / "reports"
ALLOWED_TYPES = ["xlsx", "xlsm", "xls", "csv"]


def setup_page() -> None:
    st.set_page_config(
        page_title="Laptop Daily Report",
        page_icon=None,
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown(
        """
        <style>
            :root {
                --navy: #0b2f66;
                --blue: #2f6fed;
                --ink: #000000;
                --muted: #475569;
                --line: #cfe0f5;
                --soft: #f4f9ff;
                --page: #eaf4ff;
                --card: #ffffff;
                --button: #0b2f66;
                --button-hover: #123f84;
                --accent-soft: #dcecff;
                --green: #0f8f5f;
            }

            .stApp {
                background:
                    radial-gradient(circle at top left, rgba(47, 111, 237, 0.14), transparent 34%),
                    linear-gradient(180deg, #eaf4ff 0%, #f5fbff 46%, #ffffff 100%);
                color: var(--ink);
                font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            }

            .block-container {
                padding-top: 1rem;
                padding-bottom: 2.2rem;
                max-width: 1380px;
            }

            .topbar {
                display: flex;
                align-items: center;
                justify-content: space-between;
                margin-bottom: 18px;
                color: var(--navy);
            }

            .brand-lockup {
                display: flex;
                align-items: center;
                gap: 10px;
                font-weight: 850;
                font-size: 18px;
            }

            .brand-mark {
                width: 32px;
                height: 32px;
                border-radius: 7px;
                background: var(--button);
                color: #ffffff;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-weight: 950;
                box-shadow: 0 6px 16px rgba(11, 47, 102, 0.18);
            }

            .topbar-status {
                color: var(--muted);
                font-size: 13px;
                font-weight: 650;
                background: rgba(255, 255, 255, 0.78);
                border: 1px solid var(--line);
                border-radius: 999px;
                padding: 8px 12px;
            }

            .hero {
                display: grid;
                grid-template-columns: minmax(0, 1.05fr) minmax(360px, 0.95fr);
                gap: 24px;
                align-items: stretch;
                margin-bottom: 22px;
            }

            .hero-copy {
                background: rgba(255, 255, 255, 0.9);
                border: 1px solid var(--line);
                border-radius: 8px;
                padding: 30px;
                box-shadow: 0 18px 42px rgba(11, 47, 102, 0.08);
            }

            .eyebrow {
                display: inline-flex;
                align-items: center;
                color: #000000;
                background: var(--accent-soft);
                border: 1px solid #b8d4f8;
                border-radius: 999px;
                padding: 7px 11px;
                font-size: 12px;
                line-height: 1;
                font-weight: 850;
                margin-bottom: 16px;
            }

            .hero-copy h1 {
                color: #000000;
                font-size: 42px;
                line-height: 1.07;
                margin: 0 0 14px 0;
                letter-spacing: 0;
                font-weight: 900;
            }

            .hero-copy p {
                margin: 0;
                color: var(--muted);
                font-size: 16px;
                line-height: 1.55;
                max-width: 620px;
            }

            .hero-actions {
                display: flex;
                gap: 12px;
                flex-wrap: wrap;
                margin-top: 22px;
            }

            .action-chip {
                border: 1px solid var(--line);
                background: #f4f9ff;
                border-radius: 7px;
                padding: 10px 13px;
                color: #000000;
                font-size: 13px;
                font-weight: 750;
            }

            .sheet-visual {
                background: rgba(255, 255, 255, 0.92);
                border-radius: 8px;
                padding: 18px;
                min-height: 282px;
                color: var(--ink);
                box-shadow: 0 18px 42px rgba(11, 47, 102, 0.08);
                border: 1px solid var(--line);
            }

            .sheet-toolbar {
                display: flex;
                align-items: center;
                justify-content: space-between;
                margin-bottom: 14px;
                color: var(--navy);
                font-size: 12px;
                font-weight: 750;
            }

            .toolbar-dots {
                display: flex;
                gap: 6px;
            }

            .toolbar-dot {
                width: 9px;
                height: 9px;
                border-radius: 50%;
                background: var(--blue);
                opacity: .92;
            }

            .sheet-grid {
                background: white;
                border-radius: 7px;
                overflow: hidden;
                border: 1px solid var(--line);
            }

            .sheet-row {
                display: grid;
                grid-template-columns: 1.1fr .75fr .75fr .75fr;
                min-height: 36px;
                border-bottom: 1px solid #e7edf4;
            }

            .sheet-row:last-child {
                border-bottom: none;
            }

            .sheet-cell {
                padding: 10px 12px;
                color: var(--ink);
                border-right: 1px solid #e7edf4;
                font-size: 12px;
                font-weight: 650;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }

            .sheet-cell:last-child {
                border-right: none;
            }

            .sheet-head .sheet-cell {
                background: #eff4ff;
                color: var(--navy);
                font-weight: 850;
            }

            .sheet-total .sheet-cell {
                background: #e6f1ff;
                color: #000000;
                font-weight: 900;
            }

            .sheet-accent {
                color: var(--green);
            }

            .section-title {
                font-size: 16px;
                font-weight: 850;
                color: #000000;
                margin: 12px 0 8px;
            }

            div[data-testid="stMetric"] {
                background: rgba(255, 255, 255, 0.9);
                border: 1px solid var(--line);
                border-radius: 8px;
                padding: 12px 14px;
                box-shadow: 0 10px 28px rgba(11, 47, 102, 0.06);
                min-height: 86px;
                overflow: hidden;
            }

            div[data-testid="stMetric"] label {
                color: var(--muted);
                font-weight: 750;
            }

            div[data-testid="stMetric"] [data-testid="stMetricValue"] {
                font-size: 24px;
                line-height: 1.15;
                color: #000000;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }

            div[data-testid="stMetric"] [data-testid="stMetricLabel"] {
                font-size: 12px;
                line-height: 1.2;
                white-space: nowrap;
            }

            div[data-testid="stFileUploader"] section {
                border: 1px dashed #9fc2ee;
                background: #ffffff;
                border-radius: 8px;
            }

            div[data-testid="stFileUploader"] * {
                color: var(--ink);
            }

            div[data-testid="stFileUploader"] section:hover {
                border-color: var(--blue);
                background: #f4f9ff;
            }

            div[data-testid="stTabs"] button {
                font-weight: 850;
                color: #000000 !important;
            }

            div[data-testid="stTabs"] button * {
                color: #000000;
            }

            div[data-testid="stTabs"] button[role="tab"],
            div[data-testid="stTabs"] button[role="tab"] p,
            div[data-testid="stTabs"] button[role="tab"] span,
            button[data-baseweb="tab"],
            button[data-baseweb="tab"] p,
            button[data-baseweb="tab"] span {
                color: #000000 !important;
            }

            div[data-testid="stTabs"] button[aria-selected="true"],
            div[data-testid="stTabs"] button[aria-selected="true"] p,
            div[data-testid="stTabs"] button[aria-selected="true"] span,
            button[data-baseweb="tab"][aria-selected="true"],
            button[data-baseweb="tab"][aria-selected="true"] p,
            button[data-baseweb="tab"][aria-selected="true"] span {
                color: #000000 !important;
            }

            div[data-testid="stTabs"] [data-baseweb="tab-highlight"] {
                background-color: var(--button) !important;
            }

            div[data-baseweb="input"],
            div[data-baseweb="select"],
            div[data-baseweb="popover"] div,
            div[data-testid="stDateInput"] div,
            div[data-testid="stTextInput"] div {
                background-color: #ffffff !important;
                color: #000000 !important;
                border-color: #9fc2ee !important;
            }

            input,
            textarea,
            input::placeholder,
            textarea::placeholder {
                color: #000000 !important;
                background-color: #ffffff !important;
            }

            div[data-testid="stTextInput"] input,
            div[data-testid="stDateInput"] input {
                color: #000000 !important;
                background-color: #ffffff !important;
                -webkit-text-fill-color: #000000 !important;
            }

            div[data-testid="stTextInput"] label,
            div[data-testid="stDateInput"] label,
            div[data-testid="stFileUploader"] label,
            div[data-testid="stFileUploader"] p,
            div[data-testid="stFileUploader"] span,
            div[data-testid="stFileUploader"] small {
                color: #000000 !important;
            }

            .stButton button, .stDownloadButton button {
                border-radius: 6px;
                min-height: 42px;
                font-weight: 850;
                border: 1px solid var(--button);
                background: #ffffff;
                color: #000000;
            }

            .stButton button[kind="primary"] {
                background: var(--button);
                border-color: var(--button);
                color: #ffffff;
                box-shadow: 0 10px 18px rgba(11, 47, 102, 0.18);
            }

            .stButton button[kind="primary"] *,
            .stDownloadButton button * {
                color: #ffffff !important;
            }

            .stButton button[kind="primary"]:hover,
            .stDownloadButton button:hover {
                background: var(--button-hover);
                border-color: var(--button-hover);
                color: #ffffff;
            }

            .stDownloadButton button {
                background: var(--button);
                border-color: var(--button);
                color: #ffffff;
                box-shadow: 0 10px 18px rgba(11, 47, 102, 0.18);
            }

            div[data-testid="stFileUploader"] button,
            div[data-testid="stFileUploader"] button[kind],
            button[data-testid="stBaseButton-secondary"] {
                background: var(--button) !important;
                border-color: var(--button) !important;
                color: #ffffff !important;
                border-radius: 6px !important;
                font-weight: 850 !important;
            }

            div[data-testid="stFileUploader"] button *,
            button[data-testid="stBaseButton-secondary"] * {
                color: #ffffff !important;
            }

            .preview-table-wrap {
                width: 100%;
                max-height: 520px;
                overflow-x: auto;
                overflow-y: auto;
                background: #ffffff;
                border: 1px solid var(--line);
                border-radius: 8px;
                box-shadow: 0 10px 28px rgba(11, 47, 102, 0.06);
            }

            .preview-table {
                width: 100%;
                min-width: 920px;
                border-collapse: collapse;
                color: #000000;
                font-size: 13px;
            }

            .preview-table th,
            .preview-table td {
                border: 1px solid #e7edf4;
                padding: 8px 10px;
                white-space: nowrap;
            }

            .preview-table th {
                background: #eff4ff;
                color: var(--navy);
                font-weight: 850;
                position: sticky;
                top: 0;
                z-index: 1;
                text-align: center;
                box-shadow: inset 0 -1px 0 var(--line);
            }

            .preview-table td {
                text-align: right;
            }

            .preview-table th:first-child,
            .preview-table td:first-child {
                text-align: left;
                font-weight: 750;
            }

            .preview-table tbody tr:nth-child(even) {
                background: #f4f9ff;
            }

            .preview-table tbody tr:last-child {
                background: #dcecff;
                font-weight: 900;
            }

            .preview-table tbody tr:hover {
                background: #eaf4ff;
            }

            .status-card {
                background: rgba(255, 255, 255, 0.9);
                border: 1px solid var(--line);
                border-radius: 8px;
                padding: 14px 16px;
            }

            .panel-title {
                font-size: 18px;
                font-weight: 900;
                color: #000000;
                margin: 0 0 4px 0;
            }

            .panel-subtitle {
                color: var(--muted);
                font-size: 13px;
                margin-bottom: 14px;
            }

            div[data-testid="stAlert"] {
                border-radius: 8px;
            }

            @media (max-width: 860px) {
                .hero {
                    grid-template-columns: 1fr;
                }

                .hero-copy h1 {
                    font-size: 34px;
                }

                .sheet-visual {
                    min-height: 240px;
                }
            }

            /* Laptop report shell, matched to the localhost:5173 reference app. */
            :root {
                --shell-page: #f1f5f9;
                --shell-card: #ffffff;
                --shell-blue: #2563eb;
                --shell-blue-hover: #1d4ed8;
                --shell-text: #1e293b;
                --shell-muted: #64748b;
                --shell-line: #e2e8f0;
            }

            header[data-testid="stHeader"],
            div[data-testid="stToolbar"],
            #MainMenu,
            footer {
                display: none !important;
            }

            .stApp {
                background: var(--shell-page) !important;
                color: var(--shell-text);
            }

            .block-container {
                max-width: none;
                padding: 6.25rem 2rem 2rem 2rem !important;
            }

            .app-header {
                position: fixed;
                top: 0;
                left: 0;
                right: 0;
                z-index: 999;
                background: #ffffff;
                border-bottom: 1px solid #e2e8f0;
                min-height: 4.75rem;
                display: flex;
                align-items: center;
                justify-content: space-between;
                padding: 1rem 1.5rem;
            }

            .app-header h1 {
                margin: 0;
                color: #1e293b;
                font-size: 1.25rem;
                line-height: 1.2;
                font-weight: 800;
                letter-spacing: 0;
            }

            .app-header-meta {
                display: flex;
                align-items: center;
                gap: 0.5rem;
                color: #64748b;
                font-size: 0.8rem;
                font-weight: 700;
            }

            .app-header-pill {
                padding: 0.38rem 0.72rem;
                border-radius: 999px;
                background: #f1f5f9;
                color: #475569;
                border: 1px solid #e2e8f0;
            }

            .button-label-spacer {
                height: 1.72rem;
            }

            div[data-testid="stVerticalBlockBorderWrapper"] {
                background: #ffffff;
                border: 1px solid #e2e8f0 !important;
                border-radius: 0.75rem !important;
                box-shadow: 0 10px 22px rgba(15, 23, 42, 0.08);
            }

            div[data-testid="stMetric"] {
                border-color: #e2e8f0 !important;
                border-radius: 0.75rem !important;
                box-shadow: 0 10px 22px rgba(15, 23, 42, 0.07) !important;
            }

            div[data-testid="stMetric"] label,
            div[data-testid="stMetric"] [data-testid="stMetricLabel"] {
                color: #64748b !important;
                font-weight: 700;
            }

            div[data-testid="stMetric"] [data-testid="stMetricValue"] {
                color: #1e293b !important;
                font-size: 1.35rem !important;
                font-weight: 850;
            }

            .section-title,
            .panel-title {
                color: #475569 !important;
                font-weight: 800;
            }

            .panel-subtitle,
            .stCaptionContainer,
            .stMarkdown p {
                color: #64748b;
            }

            div[data-baseweb="input"],
            div[data-baseweb="select"],
            div[data-testid="stDateInput"] div,
            div[data-testid="stTextInput"] div {
                border-color: #cbd5e1 !important;
                border-radius: 0.5rem !important;
                background: #ffffff !important;
                color: #1e293b !important;
            }

            div[data-testid="stTextInput"] input,
            div[data-testid="stDateInput"] input {
                color: #1e293b !important;
                -webkit-text-fill-color: #1e293b !important;
            }

            div[data-testid="stTextInput"] label,
            div[data-testid="stDateInput"] label,
            div[data-testid="stSelectbox"] label,
            div[data-testid="stFileUploader"] label {
                color: #475569 !important;
                font-weight: 750 !important;
            }

            .stButton button,
            .stDownloadButton button {
                border-radius: 0.5rem !important;
                min-height: 2.75rem !important;
                border-color: var(--shell-blue) !important;
                font-weight: 800 !important;
                transition: all 0.18s ease;
            }

            .stButton button[kind="primary"],
            .stDownloadButton button {
                background: var(--shell-blue) !important;
                color: #ffffff !important;
                box-shadow: 0 10px 20px rgba(37, 99, 235, 0.22) !important;
            }

            .stButton button[kind="primary"]:hover,
            .stDownloadButton button:hover {
                background: var(--shell-blue-hover) !important;
                border-color: var(--shell-blue-hover) !important;
                transform: translateY(-1px);
            }

            div[data-testid="stFileUploader"] section {
                border: 2px dashed #cbd5e1 !important;
                border-radius: 0.75rem !important;
                background: #ffffff !important;
                min-height: 132px;
                transition: all 0.22s ease;
            }

            div[data-testid="stFileUploader"] section:hover {
                border-color: #60a5fa !important;
                background: #eff6ff !important;
            }

            div[data-testid="stFileUploader"] button,
            button[data-testid="stBaseButton-secondary"] {
                border-radius: 0.5rem !important;
                background: #2563eb !important;
                border-color: #2563eb !important;
                color: #ffffff !important;
            }

            .preview-table-wrap {
                border-color: #e2e8f0 !important;
                border-radius: 0.75rem !important;
                box-shadow: 0 10px 22px rgba(15, 23, 42, 0.08) !important;
            }

            .preview-table {
                color: #1e293b !important;
            }

            .preview-table th {
                background: #f8fafc !important;
                color: #475569 !important;
                border-color: #e2e8f0 !important;
            }

            .preview-table td {
                border-color: #e2e8f0 !important;
            }

            .preview-table tbody tr:nth-child(even) {
                background: #f8fafc !important;
            }

            .preview-table tbody tr:last-child {
                background: #eff6ff !important;
                color: #1d4ed8;
            }

            .status-card {
                border-color: #e2e8f0 !important;
                border-radius: 0.75rem !important;
                background: #ffffff !important;
                box-shadow: 0 10px 22px rgba(15, 23, 42, 0.06);
                color: #64748b;
            }

            @media (max-width: 980px) {
                .app-header {
                    left: 0;
                    top: auto;
                    position: sticky;
                }

                .block-container {
                    padding: 1rem !important;
                }
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_app_chrome() -> None:
    st.markdown(
        """
        <header class="app-header">
            <h1>Laptop Category Report</h1>
            <div class="app-header-meta">
                <span class="app-header-pill">LAPTOP</span>
                <span class="app-header-pill">Local App</span>
            </div>
        </header>
        """,
        unsafe_allow_html=True,
    )


def init_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def clean_name(name: str) -> str:
    allowed = []
    for char in Path(name).name:
        allowed.append(char if char.isalnum() or char in "._- " else "_")
    cleaned = "".join(allowed).strip()
    return cleaned or "uploaded_file"


def save_uploaded_file(uploaded_file, prefix: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = UPLOAD_DIR / f"{prefix}_{timestamp}_{clean_name(uploaded_file.name)}"
    path.write_bytes(uploaded_file.getbuffer())
    return path


def db_status() -> tuple[bool, str]:
    try:
        ensure_schema()
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
            return True, "Connected"
        finally:
            conn.close()
    except Exception as exc:
        return False, str(exc)


def latest_sales_date() -> date:
    try:
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT MAX(sale_date) FROM fact_sales")
                value = cur.fetchone()[0]
                return value or date.today()
        finally:
            conn.close()
    except Exception:
        return date.today()


def latest_stock_date() -> date | None:
    try:
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT MAX(stock_date) FROM fact_stock")
                return cur.fetchone()[0]
        finally:
            conn.close()
    except Exception:
        return None


def stock_date_exists(stock_date: date) -> bool:
    try:
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT EXISTS (SELECT 1 FROM fact_stock WHERE stock_date = %s)",
                    (stock_date,),
                )
                return bool(cur.fetchone()[0])
        finally:
            conn.close()
    except Exception:
        return False


def show_database_status() -> None:
    ok, message = db_status()
    latest_date = latest_sales_date() if ok else None
    stock_date = latest_stock_date() if ok else None
    col1, col2, col3, col4 = st.columns([1, 1, 1.25, 1.25])
    with col1:
        st.metric("Database", "Online" if ok else "Offline")
    with col2:
        st.metric("Category", "LAPTOP")
    with col3:
        st.metric("Latest Sales", latest_date.strftime("%Y-%m-%d") if latest_date else "-")
    with col4:
        st.metric("Latest Stock", stock_date.strftime("%Y-%m-%d") if stock_date else "-")
    if ok:
        st.success("PostgreSQL connection is ready.")
    else:
        st.error(f"Database connection failed: {message}")


def sales_tab(uploaded_by: str) -> None:
    st.markdown('<div class="section-title">Revenue Upload</div>', unsafe_allow_html=True)
    sales_file = st.file_uploader("REVENUEERA file", type=ALLOWED_TYPES, key="sales_file")

    upload_clicked = st.button("Upload Revenue", type="primary", use_container_width=True)

    if not sales_file and upload_clicked:
        st.warning("Choose a REVENUEERA file first.")
        return

    if sales_file and upload_clicked:
        path = save_uploaded_file(sales_file, "sales")
        with st.spinner("Uploading sales data..."):
            try:
                inserted, updated = upload_sales(path, uploaded_by)
                st.success(f"Sales uploaded. Inserted: {inserted:,}. Updated: {updated:,}.")
            except Exception as exc:
                st.error(f"Sales upload failed: {exc}")


def stock_tab(uploaded_by: str, default_stock_date: date) -> None:
    st.markdown('<div class="section-title">Stock Upload</div>', unsafe_allow_html=True)
    stock_file = st.file_uploader("GeneralInventory file", type=ALLOWED_TYPES, key="stock_file")

    detected_stock_date = (
        stock_date_from_filename(stock_file.name, default_stock_date)
        if stock_file
        else default_stock_date
    )
    date_key = f"stock_date_{clean_name(stock_file.name)}" if stock_file else "stock_date_empty"
    stock_date = st.date_input(
        "Stock date",
        value=detected_stock_date,
        format="YYYY-MM-DD",
        key=date_key,
    )
    if stock_file and detected_stock_date != default_stock_date:
        st.caption(f"Detected from filename: {detected_stock_date:%Y-%m-%d}")

    upload_clicked = st.button("Upload Stock", type="primary", use_container_width=True)

    if not stock_file and upload_clicked:
        st.warning("Choose a GeneralInventory file first.")
        return

    if stock_file and upload_clicked:
        path = save_uploaded_file(stock_file, "stock")
        with st.spinner("Uploading stock data..."):
            try:
                inserted, updated = upload_stock(path, uploaded_by, stock_date)
                st.success(f"Stock uploaded. Inserted: {inserted:,}. Updated: {updated:,}.")
            except Exception as exc:
                st.error(f"Stock upload failed: {exc}")


def report_tab(report_date: date) -> None:
    st.markdown('<div class="section-title">Excel Report</div>', unsafe_allow_html=True)
    output_name = f"LAPTOP_REPORT_{report_date:%Y%m%d}.xlsx"
    output_path = REPORT_DIR / output_name

    if st.button("Generate Excel Report", type="primary", use_container_width=True):
        with st.spinner("Generating Excel report..."):
            try:
                generate_report(report_date, output_path)
                st.session_state["latest_report"] = str(output_path)
                st.success(f"Report generated: {output_name}")
            except Exception as exc:
                st.error(f"Report generation failed: {exc}")

    latest = st.session_state.get("latest_report")
    if latest and Path(latest).exists():
        path = Path(latest)
        st.download_button(
            label="Download Excel Report",
            data=path.read_bytes(),
            file_name=path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.markdown('<div class="status-card">No report generated in this session.</div>', unsafe_allow_html=True)


def generated_report_files() -> list[Path]:
    if not REPORT_DIR.exists():
        return []
    return sorted(REPORT_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)


def preview_report_data(path: Path, sheet_name: str, max_rows: int | None = None) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} tidak ditemukan.")

        ws = wb[sheet_name]
        sheet_rows = ws.iter_rows(values_only=True)
        next(sheet_rows, ())
        group_row = next(sheet_rows, ())
        header_row = next(sheet_rows, ())
        if sheet_name == "SHOP LEVEL":
            headers = []
            current_group = ""
            for col, subheader in enumerate(header_row, start=1):
                group = group_row[col - 1] if col <= len(group_row) else None
                if group:
                    current_group = str(group).strip()
                if col == 1:
                    header = subheader or "STORE NAME"
                else:
                    header = f"{current_group} {subheader or ''}".strip()
                headers.append(header or f"Column {col}")
        else:
            headers = list(header_row)
            headers = [header if header else f"Column {idx}" for idx, header in enumerate(headers, start=1)]

        rows = []
        for row_idx, values in enumerate(sheet_rows):
            if max_rows is not None and row_idx >= max_rows:
                break
            row = list(values[:len(headers)])
            if len(row) < len(headers):
                row.extend([None] * (len(headers) - len(row)))
            if any(value is not None for value in row):
                rows.append(row)
        return pd.DataFrame(rows, columns=headers)
    finally:
        wb.close()


def coerce_preview_number(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, Number) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        normalized = cleaned.replace(".", "").replace(",", ".") if "," in cleaned else cleaned.replace(",", "")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def format_id_number(value: object) -> str:
    number = coerce_preview_number(value)
    if number is None:
        return "" if value is None or pd.isna(value) else str(value)
    return f"{number:,.0f}".replace(",", ".")


def format_preview_report_data(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()
    for column in formatted.columns:
        column_name = str(column).upper()
        if column_name == "GROWTH RATE":
            formatted[column] = formatted[column].apply(
                lambda value: "" if coerce_preview_number(value) is None else f"{coerce_preview_number(value) * 100:.2f}%".replace(".", ",")
            )
        elif column_name not in {"BRAND", "STORE NAME"}:
            formatted[column] = formatted[column].apply(format_id_number)
        else:
            formatted[column] = formatted[column].fillna("")
    return formatted


def render_report_preview(path: Path, sheet_name: str) -> None:
    preview_df = preview_report_data(path, sheet_name)
    preview_df = format_preview_report_data(preview_df)
    preview_html = preview_df.to_html(
        index=False,
        border=0,
        classes="preview-table",
        escape=True,
    )
    st.markdown(f'<div class="preview-table-wrap">{preview_html}</div>', unsafe_allow_html=True)


def report_number(value: object) -> float:
    number = coerce_preview_number(value)
    return number if number is not None else 0.0


def prepare_visualization_data(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.Series]:
    laptop_df = preview_report_data(path, "LAPTOP REPORT")
    shop_df = preview_report_data(path, "SHOP LEVEL")

    total_rows = laptop_df[laptop_df["BRAND"].astype(str).str.upper() == "GRAND TOTAL"]
    total = total_rows.iloc[0] if not total_rows.empty else laptop_df.sum(numeric_only=True)
    laptop_df = laptop_df[laptop_df["BRAND"].astype(str).str.upper() != "GRAND TOTAL"].copy()
    shop_df = shop_df[shop_df["STORE NAME"].astype(str).str.upper() != "GRAND TOTAL"].copy()

    for column in laptop_df.columns:
        if column != "BRAND":
            laptop_df[column] = laptop_df[column].apply(report_number)
    for column in shop_df.columns:
        if column != "STORE NAME":
            shop_df[column] = shop_df[column].apply(report_number)
    return laptop_df, shop_df, total


def style_chart(chart: alt.Chart) -> alt.Chart:
    return (
        chart.configure_view(stroke=None)
        .configure_axis(
            labelColor="#475569",
            titleColor="#64748b",
            gridColor="#e2e8f0",
            domainColor="#cbd5e1",
            tickColor="#cbd5e1",
            labelFontSize=11,
            titleFontSize=11,
        )
        .configure_title(
            color="#1e293b",
            fontSize=15,
            fontWeight=700,
            anchor="start",
            offset=14,
        )
        .configure_legend(
            labelColor="#475569",
            titleColor="#475569",
            orient="top",
            direction="horizontal",
        )
    )


def render_visualization(path: Path) -> None:
    laptop_df, shop_df, total = prepare_visualization_data(path)
    if laptop_df.empty:
        st.info("Belum ada data yang dapat divisualisasikan.")
        return

    day_revenue = report_number(total.get("DAY Rev (IDR)"))
    mtd_revenue = report_number(total.get("MTD Rev (IDR)"))
    mtd_qty = report_number(total.get("MTD Qty"))
    stock_units = (
        report_number(total.get("New"))
        + report_number(total.get("Error (New)"))
        + report_number(total.get("Display"))
    )
    growth_rate = report_number(total.get("GROWTH RATE")) * 100

    metric_cols = st.columns(4)
    metric_cols[0].metric("Day Revenue", f"Rp {format_id_number(day_revenue)}")
    metric_cols[1].metric(
        "MTD Revenue",
        f"Rp {format_id_number(mtd_revenue)}",
        delta=f"{growth_rate:.2f}% vs same period".replace(".", ","),
    )
    metric_cols[2].metric("MTD Quantity", format_id_number(mtd_qty))
    metric_cols[3].metric("Stock Units", format_id_number(stock_units))

    revenue_data = laptop_df[["BRAND", "MTD Rev (IDR)"]].copy()
    revenue_data["Revenue Label"] = revenue_data["MTD Rev (IDR)"].apply(
        lambda value: f"Rp {format_id_number(value)}"
    )
    revenue_chart = (
        alt.Chart(revenue_data)
        .mark_bar(cornerRadiusEnd=4, color="#2563eb")
        .encode(
            x=alt.X("MTD Rev (IDR):Q", title="MTD Revenue (IDR)", axis=alt.Axis(format="~s")),
            y=alt.Y("BRAND:N", title=None, sort="-x"),
            tooltip=[
                alt.Tooltip("BRAND:N", title="Brand"),
                alt.Tooltip("Revenue Label:N", title="MTD Revenue"),
            ],
        )
        .properties(title="MTD Revenue by Brand", height=max(260, len(revenue_data) * 31))
    )

    stock_data = laptop_df[["BRAND", "New", "Error (New)", "Display"]].melt(
        id_vars="BRAND",
        value_vars=["New", "Error (New)", "Display"],
        var_name="Stock Type",
        value_name="Units",
    )
    stock_chart = (
        alt.Chart(stock_data)
        .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
        .encode(
            x=alt.X("BRAND:N", title=None, sort=alt.EncodingSortField(field="Units", op="sum", order="descending")),
            y=alt.Y("Units:Q", title="Units", axis=alt.Axis(format="~s")),
            color=alt.Color(
                "Stock Type:N",
                title=None,
                scale=alt.Scale(
                    domain=["New", "Error (New)", "Display"],
                    range=["#0f8f5f", "#dc2626", "#f59e0b"],
                ),
            ),
            tooltip=[
                alt.Tooltip("BRAND:N", title="Brand"),
                alt.Tooltip("Stock Type:N", title="Stock"),
                alt.Tooltip("Units:Q", title="Units", format=",.0f"),
            ],
        )
        .properties(title="Stock Composition", height=260)
    )

    growth_data = laptop_df[["BRAND", "GROWTH RATE"]].copy()
    growth_data["Growth Percent"] = growth_data["GROWTH RATE"] * 100
    growth_chart = (
        alt.Chart(growth_data)
        .mark_bar(cornerRadiusEnd=4)
        .encode(
            x=alt.X("Growth Percent:Q", title="Growth vs Same Period", axis=alt.Axis(format=".1f")),
            y=alt.Y("BRAND:N", title=None, sort="-x"),
            color=alt.condition(
                alt.datum["Growth Percent"] >= 0,
                alt.value("#0f8f5f"),
                alt.value("#dc2626"),
            ),
            tooltip=[
                alt.Tooltip("BRAND:N", title="Brand"),
                alt.Tooltip("Growth Percent:Q", title="Growth (%)", format=".2f"),
            ],
        )
        .properties(title="Brand Growth Rate", height=max(260, len(growth_data) * 31))
    )

    if shop_df.empty:
        top_shop_data = pd.DataFrame(columns=["STORE NAME", "TOTAL MTD Rev"])
    else:
        top_shop_data = (
            shop_df[["STORE NAME", "TOTAL MTD Rev"]]
            .nlargest(15, "TOTAL MTD Rev")
            .sort_values("TOTAL MTD Rev", ascending=True)
            .copy()
        )
    top_shop_data["Revenue Label"] = top_shop_data["TOTAL MTD Rev"].apply(
        lambda value: f"Rp {format_id_number(value)}"
    )
    shop_chart = (
        alt.Chart(top_shop_data)
        .mark_bar(cornerRadiusEnd=4, color="#0891b2")
        .encode(
            x=alt.X("TOTAL MTD Rev:Q", title="MTD Revenue (IDR)", axis=alt.Axis(format="~s")),
            y=alt.Y("STORE NAME:N", title=None, sort=None),
            tooltip=[
                alt.Tooltip("STORE NAME:N", title="Store"),
                alt.Tooltip("Revenue Label:N", title="MTD Revenue"),
            ],
        )
        .properties(title="Top 15 Stores by MTD Revenue", height=420)
    )

    top_left, top_right = st.columns(2, gap="large")
    with top_left:
        st.altair_chart(style_chart(revenue_chart), width="stretch")
    with top_right:
        st.altair_chart(style_chart(stock_chart), width="stretch")

    bottom_left, bottom_right = st.columns(2, gap="large")
    with bottom_left:
        st.altair_chart(style_chart(growth_chart), width="stretch")
    with bottom_right:
        if top_shop_data.empty:
            st.info("Data Shop Level belum tersedia.")
        else:
            st.altair_chart(style_chart(shop_chart), width="stretch")


def generated_data_panel(default_sales_date: date, default_stock_date: date) -> None:
    st.markdown('<div class="section-title">Generated Data</div>', unsafe_allow_html=True)

    with st.container(border=True):
        sales_date_col, stock_date_col, action_col = st.columns([1, 1, 1])
        with sales_date_col:
            report_date = st.date_input("Sales date", value=default_sales_date, format="YYYY-MM-DD")
        with stock_date_col:
            report_stock_date = st.date_input(
                "Stock date",
                value=default_stock_date,
                format="YYYY-MM-DD",
                key="report_stock_date",
            )
        with action_col:
            st.markdown('<div class="button-label-spacer"></div>', unsafe_allow_html=True)
            generate_clicked = st.button("Generate Excel", type="primary", use_container_width=True)

        output_name = f"LAPTOP_REPORT_{report_date:%Y%m%d}.xlsx"
        output_path = REPORT_DIR / output_name
        status_message: tuple[str, str] | None = None

        if generate_clicked:
            if not stock_date_exists(report_stock_date):
                status_message = (
                    "error",
                    f"Stock data for {report_stock_date:%Y-%m-%d} is not available. Upload GeneralInventory for that date first.",
                )
            else:
                with st.spinner("Generating Excel report..."):
                    try:
                        generate_report(report_date, output_path, report_stock_date)
                        st.session_state["latest_report"] = str(output_path)
                        status_message = (
                            "success",
                            f"Report generated: {output_name} | Stock: {report_stock_date:%Y-%m-%d}",
                        )
                    except Exception as exc:
                        status_message = ("error", f"Report generation failed: {exc}")

        files = generated_report_files()
        if not files:
            st.markdown(
                '<div class="status-card">Belum ada report yang dibuat. Upload file lalu klik Generate Excel.</div>',
                unsafe_allow_html=True,
            )
            return

        latest_session = st.session_state.get("latest_report")
        selected = files[0]
        if latest_session:
            session_path = Path(latest_session)
            if session_path.exists():
                selected = session_path

        _, _, download_col = st.columns([1, 1, 1])
        with download_col:
            st.download_button(
                label="Download Excel",
                data=selected.read_bytes(),
                file_name=selected.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        if status_message:
            level, message = status_message
            if level == "success":
                st.success(message)
            else:
                st.error(message)

        laptop_tab, shop_tab, visualization_tab = st.tabs(["Laptop Report", "Shop Level", "Visualization"])
        with laptop_tab:
            try:
                render_report_preview(selected, "LAPTOP REPORT")
            except Exception as exc:
                st.warning(f"Preview Laptop Report tidak tersedia: {exc}")
        with shop_tab:
            try:
                render_report_preview(selected, "SHOP LEVEL")
            except Exception as exc:
                st.warning(f"Preview Shop Level tidak tersedia: {exc}")
        with visualization_tab:
            try:
                render_visualization(selected)
            except Exception as exc:
                st.warning(f"Visualisasi tidak tersedia: {exc}")


def upload_workflow_panel(uploaded_by: str, default_stock_date: date) -> None:
    with st.container(border=True):
        st.markdown('<div class="panel-title">Upload Data</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="panel-subtitle">Upload REVENUEERA and GeneralInventory in one place before generating the report.</div>',
            unsafe_allow_html=True,
        )
        sales_tab(uploaded_by)
        st.divider()
        stock_tab(uploaded_by, default_stock_date)


def main() -> None:
    init_dirs()
    setup_page()
    render_app_chrome()

    default_sales_date = latest_sales_date()
    default_stock_date = latest_stock_date() or default_sales_date
    uploaded_by = "Yusuf"

    show_database_status()
    st.write("")

    upload_col, generated_col = st.columns([0.78, 1.62], gap="large")
    with upload_col:
        upload_workflow_panel(uploaded_by, default_sales_date)
    with generated_col:
        generated_data_panel(default_sales_date, default_stock_date)


if __name__ == "__main__":
    main()
